from openpyxl import Workbook
from openpyxl.styles import PatternFill


def export_to_excel(data, OUTPUT_XLSX):

    RED = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    YELLOW = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")

    # ---------- COLLECT SUBJECT-WISE DATA ----------
    subjects = {}

    for student in data:
        info = student.get("student_info", {})
        roll = info.get("Roll No")
        name = info.get("Student Name")

        for sub in student.get("attendance", {}).get("subjects", []):
            subject_name = sub["subject_name"]

            total_lectures = int(sub["total_lectures"])
            attended_lectures = int(sub["attended_lectures"])
            percent = float(sub["attendance_percentage"])

            subjects.setdefault(subject_name, []).append(
                (roll, name, total_lectures, attended_lectures, percent)
            )

    # ---------- CREATE EXCEL ----------
    wb = Workbook()
    wb.remove(wb.active)

    for subject, records in subjects.items():
        ws = wb.create_sheet(title=subject[:31])

        ws.append([
            "Roll No",
            "Student Name",
            "Total Lectures",
            "Lectures Attended",
            "Attendance %"
        ])

        for roll, name, total, attended, percent in records:
            ws.append([roll, name, total, attended, percent])

            percent_cell = ws.cell(row=ws.max_row, column=5)

            if percent < 60:
                percent_cell.fill = RED
            elif 60 <= percent < 75:
                percent_cell.fill = YELLOW
            # exactly 75 or >75 → no color

        # ---------- AUTO ADJUST COLUMN WIDTH ----------
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_length + 4

    # ---------- SAVE ----------
    wb.save(OUTPUT_XLSX)
